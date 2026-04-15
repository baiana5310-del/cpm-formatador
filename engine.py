import os
import re
import math
import unicodedata
import logging
import traceback
import shutil
from pathlib import Path
from decimal import Decimal, InvalidOperation
from datetime import date, datetime, time

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.comments import Comment

ARQUIVO_LOG = "formatador_turbo_erros.log"
logging.basicConfig(
    filename=ARQUIVO_LOG,
    level=logging.INFO,
    format="%(asctime)s - [%(levelname)s] - %(funcName)s - %(message)s",
    datefmt="%d/%m/%Y %H:%M:%S"
)

# ─────────────────────────────────────────────────────────────────────────────
#  SISTEMA DE TEMAS — REDESENHADO PARA IMPACTO MÁXIMO
#  Filosofia: cada tema é uma identidade completa, não apenas uma cor.
#  Inspiração: dashboards de SaaS premium, relatórios de consultoria top-tier.
# ─────────────────────────────────────────────────────────────────────────────

TEMAS = {
    # ── SLATE EXECUTIVE ─────────────────────────────────────────────────────
    # Tom âncora: azul-ardósia profundo. Transmite autoridade silenciosa.
    # Acento: índigo brilhante. Toca de sofisticação técnica.
    "🔵 Azul Executivo": {
        # Cabeçalho — fundo quase-preto com toque azul para profundidade
        "header_bg":       "0D1B2A",
        "header_fg":       "F0F4FF",
        # Linha alternada — quase branco com frescor azulado
        "accent_light":    "F7F9FC",
        # Linha branca — branco puro (contraste limpo)
        "row_white":       "FFFFFF",
        # Linha de totais — navy intenso
        "totals_bg":       "0D1B2A",
        "totals_fg":       "FFFFFF",
        # Bordas — cinza azulado muito sutil
        "border":          "DDE3ED",
        "border_strong":   "0D1B2A",
        # Linha zebra — toque de azul gelado
        "zebra":           "F0F4FB",
        # Títulos nos painéis
        "title_fg":        "0D1B2A",
        # Tab da aba
        "tab_color":       "1A3A5C",
        # Acento decorativo — índigo vibrante
        "accent_color":    "1D4ED8",
        # Subtítulos / labels nos painéis
        "label_fg":        "64748B",
        # Fundo do painel Dashboard
        "panel_bg":        "F0F4FB",
        # Borda do card no Dashboard
        "card_border":     "C7D4E8",
        # Número/métrica nos cards
        "metric_fg":       "0D1B2A",
        # Cabeçalho sub linha (linha de grupo)
        "subheader_bg":    "1E3A5F",
        "subheader_fg":    "E8EFF8",
        # UI
        "ui_color":        "#1D4ED8",
        "ui_hover":        "#1E3A8A",
    },

    # ── OBSIDIAN MINIMAL ────────────────────────────────────────────────────
    # Fundo quase-carvão. Tipografia branca crua. Zero decoração. Máxima clareza.
    "⚫ Grafite Minimalista": {
        "header_bg":       "111827",
        "header_fg":       "F9FAFB",
        "accent_light":    "F9FAFB",
        "row_white":       "FFFFFF",
        "totals_bg":       "030712",
        "totals_fg":       "FFFFFF",
        "border":          "E5E7EB",
        "border_strong":   "111827",
        "zebra":           "F3F4F6",
        "title_fg":        "111827",
        "tab_color":       "1F2937",
        "accent_color":    "374151",
        "label_fg":        "6B7280",
        "panel_bg":        "F3F4F6",
        "card_border":     "D1D5DB",
        "metric_fg":       "111827",
        "subheader_bg":    "1F2937",
        "subheader_fg":    "F3F4F6",
        "ui_color":        "#374151",
        "ui_hover":        "#1F2937",
    },

    # ── EMERALD PRESTIGE ────────────────────────────────────────────────────
    # Verde-floresta elegante. Ativa sensação de crescimento, solidez financeira.
    "🟢 Verde Safira": {
        "header_bg":       "052E16",
        "header_fg":       "F0FDF4",
        "accent_light":    "F0FDF4",
        "row_white":       "FFFFFF",
        "totals_bg":       "052E16",
        "totals_fg":       "FFFFFF",
        "border":          "D1FAE5",
        "border_strong":   "052E16",
        "zebra":           "ECFDF5",
        "title_fg":        "052E16",
        "tab_color":       "065F46",
        "accent_color":    "059669",
        "label_fg":        "4B5563",
        "panel_bg":        "ECFDF5",
        "card_border":     "A7F3D0",
        "metric_fg":       "052E16",
        "subheader_bg":    "064E3B",
        "subheader_fg":    "D1FAE5",
        "ui_color":        "#059669",
        "ui_hover":        "#047857",
    },

    # ── AMETHYST DEEP TECH ──────────────────────────────────────────────────
    # Roxo ultradeep. Sensação de tecnologia futurista com sofisticação.
    "🟣 Roxo Deep Tech": {
        "header_bg":       "1E0938",
        "header_fg":       "F5F3FF",
        "accent_light":    "FAF5FF",
        "row_white":       "FFFFFF",
        "totals_bg":       "0D0621",
        "totals_fg":       "FFFFFF",
        "border":          "EDE9FE",
        "border_strong":   "1E0938",
        "zebra":           "F5F3FF",
        "title_fg":        "1E0938",
        "tab_color":       "4C1D95",
        "accent_color":    "7C3AED",
        "label_fg":        "6D28D9",
        "panel_bg":        "F5F3FF",
        "card_border":     "DDD6FE",
        "metric_fg":       "1E0938",
        "subheader_bg":    "2E1065",
        "subheader_fg":    "EDE9FE",
        "ui_color":        "#7C3AED",
        "ui_hover":        "#6D28D9",
    },

    # ── COGNAC HERITAGE ─────────────────────────────────────────────────────
    # Marrom mogno + creme. Ativado para relatórios executivos tradicionais.
    "🟤 Marrom Heritage": {
        "header_bg":       "1C0A00",
        "header_fg":       "FEF9F0",
        "accent_light":    "FEFCE8",
        "row_white":       "FFFFFF",
        "totals_bg":       "0E0500",
        "totals_fg":       "FEF9F0",
        "border":          "F3E8C0",
        "border_strong":   "1C0A00",
        "zebra":           "FDFAF0",
        "title_fg":        "1C0A00",
        "tab_color":       "44200E",
        "accent_color":    "92400E",
        "label_fg":        "78350F",
        "panel_bg":        "FEFCE8",
        "card_border":     "FDE68A",
        "metric_fg":       "1C0A00",
        "subheader_bg":    "3B1506",
        "subheader_fg":    "FEF3C7",
        "ui_color":        "#92400E",
        "ui_hover":        "#78350F",
    },
}


# ─────────────────────────────────────────────────────────────────────────────
#  UTILITÁRIOS DE ESTILO — Funções de factory refinadas
# ─────────────────────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=10, color="000000", italic=False, name="Calibri"):
    """
    Fonte padrão: Calibri (universal, limpa, disponível em qualquer Windows/Mac).
    Para cabeçalhos: Calibri Bold size 11 dá presença sem peso excessivo.
    """
    return Font(name=name, bold=bold, size=size, color=color, italic=italic)


def _border(color="E5E7EB", style="thin") -> Border:
    s = Side(border_style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _border_full(left_color, right_color, top_color, bottom_color,
                 left_style="thin", right_style="thin",
                 top_style="thin", bottom_style="thin") -> Border:
    return Border(
        left=Side(border_style=left_style, color=left_color),
        right=Side(border_style=right_style, color=right_color),
        top=Side(border_style=top_style, color=top_color),
        bottom=Side(border_style=bottom_style, color=bottom_color),
    )


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _align_indent(h="left", v="center", wrap=False, indent=1) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)


# ─────────────────────────────────────────────────────────────────────────────
#  UTILITÁRIOS DE TEXTO / NORMALIZAÇÃO (preservados sem alteração)
# ─────────────────────────────────────────────────────────────────────────────

def _remover_acentos(txt: str) -> str:
    return ''.join(ch for ch in unicodedata.normalize('NFKD', str(txt)) if not unicodedata.combining(ch))


def normalizar_nome_coluna(nome):
    s = str(nome or "").strip().lower()
    s = _remover_acentos(s)
    s = s.replace("%", " percentual ")
    s = re.sub(r"[\(\)\[\]\{\}\.\-\/\\]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = s.replace(" ", "_")
    return s


def valor_vazio(valor):
    return valor is None or str(valor).strip() == ""


def apenas_digitos(valor):
    return ''.join(ch for ch in str(valor) if ch.isdigit())


def normalizar_texto(valor):
    s = str(valor or "").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def normalizar_texto_upper_sem_acentos(valor):
    return _remover_acentos(normalizar_texto(valor)).upper()


def normalizar_placa(valor):
    if valor is None or str(valor).strip() == "":
        return None
    s = str(valor).strip().upper()
    s = re.sub(r"\s+", "", s)
    return s


def validar_placa(valor):
    return True


def normalizar_status(valor):
    s = normalizar_texto_upper_sem_acentos(valor).replace(" ", "_")
    mapa = {
        "PAUSADO": "PARALISADA", "PARADA": "PARALISADA", "PARALIZADA": "PARALISADA",
        "ATIVO": "ATIVO", "ATIVA": "ATIVO", "ON": "ATIVO",
        "EM_ESTOQUE": "EM_ESTOQUE", "NO_ESTOQUE": "EM_ESTOQUE",
        "DISPONIVEL": "DISPONIVEL", "LIVRE": "DISPONIVEL"
    }
    return mapa.get(s, s)


def formatar_cpf(digitos):
    return f"{digitos[:3]}.{digitos[3:6]}.{digitos[6:9]}-{digitos[9:]}"


def formatar_cnpj(digitos):
    return f"{digitos[:2]}.{digitos[2:5]}.{digitos[5:8]}/{digitos[8:12]}-{digitos[12:]}"


def formatar_cep(digitos):
    return f"{digitos[:5]}-{digitos[5:]}"


def formatar_telefone_br(digitos):
    if len(digitos) == 10:
        return f"({digitos[:2]}) {digitos[2:6]}-{digitos[6:]}"
    if len(digitos) == 11:
        return f"({digitos[:2]}) {digitos[2:7]}-{digitos[7:]}"
    return None


def eh_email_valido(valor):
    s = normalizar_texto(valor)
    return re.fullmatch(r"^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$", s) is not None


def parse_decimal_br(valor):
    if valor_vazio(valor): return None
    if isinstance(valor, bool): return None
    if isinstance(valor, (int, float, Decimal)):
        if isinstance(valor, float) and (math.isnan(valor) or math.isinf(valor)): return None
        return Decimal(str(valor))
    s = str(valor).strip()
    s = s.replace("R$", "").replace("r$", "").replace("%", "").strip()
    s = s.replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def parse_inteiro(valor):
    dec = parse_decimal_br(valor)
    if dec is None: return None
    if dec != dec.to_integral_value(): return None
    return int(dec)


def _converter_texto_para_data(valor):
    if isinstance(valor, datetime): return valor
    if isinstance(valor, date): return datetime.combine(valor, datetime.min.time())
    if isinstance(valor, str):
        texto = valor.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(texto, fmt)
            except ValueError:
                pass
    return None


def processar_hora(valor):
    if valor_vazio(valor): return None
    if valor in (0, "0"): return None
    if isinstance(valor, time): return valor
    if isinstance(valor, datetime): return valor.time()
    if isinstance(valor, (int, float, Decimal)):
        try:
            val_float = float(valor)
            if val_float <= 0: return None
            frac = val_float % 1
            total_segundos = int(round(frac * 86400))
            h = (total_segundos // 3600) % 24
            m = (total_segundos % 3600) // 60
            s = total_segundos % 60
            return time(h, m, s)
        except (ValueError, TypeError):
            pass
    if isinstance(valor, str):
        texto = valor.strip().lower()
        formatos = ["%H:%M:%S", "%H:%M", "%Hh%M", "%H:%M:%S.%f", "%I:%M %p", "%I:%M:%S %p"]
        for fmt in formatos:
            try:
                return datetime.strptime(texto, fmt).time()
            except ValueError:
                continue
        digitos = apenas_digitos(texto)
        if len(digitos) in (3, 4):
            try:
                h = int(digitos[:-2])
                m = int(digitos[-2:])
                if 0 <= h <= 23 and 0 <= m <= 59:
                    return time(h, m)
            except ValueError:
                pass
    return None


def linha_vazia(valores):
    return all(v is None or str(v).strip() == "" for v in valores)


def copiar_linha(ws_destino, linha_destino, valores):
    for col_idx, valor in enumerate(valores, start=1):
        ws_destino.cell(row=linha_destino, column=col_idx, value=valor)


def limpar_nome_aba(nome, sheetnames):
    for ch in ['\\', '/', '*', '[', ']', ':', '?']:
        nome = nome.replace(ch, '_')
    nome = (nome.strip() or "Aba")[:31]
    base, i = nome, 1
    while nome in sheetnames:
        sufixo = f"_{i}"
        nome = f"{base[:31-len(sufixo)]}{sufixo}"
        i += 1
    return nome


def _set_cell_if_changed(cell, novo_valor):
    if cell.value != novo_valor:
        cell.value = novo_valor
        return True
    return False


def detectar_modulo_por_aba(nome_aba):
    if not nome_aba: return "DESCONHECIDO"
    nome_norm = _remover_acentos(nome_aba).lower()
    if any(k in nome_norm for k in ["obra", "projeto", "fase"]): return "OBRAS"
    if any(k in nome_norm for k in ["equipe", "func", "rh", "ponto", "presenca"]): return "EQUIPE"
    if any(k in nome_norm for k in ["veiculo", "frota", "carro", "caminhao"]): return "VEICULOS"
    if any(k in nome_norm for k in ["materiai", "material", "estoque", "almoxarifado"]): return "MATERIAIS"
    if any(k in nome_norm for k in ["entrega", "logistica", "recebimento", "expedicao"]): return "ENTREGAS"
    return "DESCONHECIDO"


def normalizar_status_por_modulo(valor, nome_aba, original):
    s_norm = normalizar_status(valor)
    s_lookup = s_norm.replace(" ", "_")
    mod = detectar_modulo_por_aba(nome_aba)

    regras = {
        "OBRAS": ["EM_ANDAMENTO", "PARALISADA", "CONCLUIDA", "CANCELADA"],
        "EQUIPE": ["ATIVO", "INATIVO", "FERIAS", "AFASTADO", "PRESENTE", "FALTA"],
        "VEICULOS": ["DISPONIVEL", "EM_USO", "MANUTENCAO", "INATIVO"],
        "MATERIAIS": ["EM_ESTOQUE", "BAIXO_ESTOQUE", "ESGOTADO", "RESERVADO"],
        "ENTREGAS": ["PENDENTE", "EM_ROTA", "ENTREGUE", "ATRASADA", "CANCELADA", "RECEBIDO", "CONFERIDO"]
    }

    mapas_contextuais = {
        "MATERIAIS": {"ACABOU": "ESGOTADO", "ZERADO": "ESGOTADO", "FALTA": "ESGOTADO"},
        "EQUIPE": {"FALTOSO": "FALTA", "AUSENTE": "FALTA"}
    }

    mapa_geral = {
        "PAUSADO": "PARALISADA", "PARADA": "PARALISADA", "PARALISADO": "PARALISADA", "PARALIZADA": "PARALISADA",
        "ANDAMENTO": "EM_ANDAMENTO", "EM_ANDAMENTO": "EM_ANDAMENTO", "EXECUTANDO": "EM_ANDAMENTO",
        "FINALIZADO": "CONCLUIDA", "FINALIZADA": "CONCLUIDA", "CONCLUIDA": "CONCLUIDA", "CONCLUIDO": "CONCLUIDA", "PRONTO": "CONCLUIDA",
        "CANCELADA": "CANCELADA", "CANCELADO": "CANCELADA",
        "ATIVO": "ATIVO", "ATIVA": "ATIVO", "TRABALHANDO": "ATIVO", "ON": "ATIVO",
        "INATIVO": "INATIVO", "INATIVA": "INATIVO", "OFF": "INATIVO",
        "FERIAS": "FERIAS", "FÉRIAS": "FERIAS",
        "AFASTADO": "AFASTADO", "AFASTADA": "AFASTADO", "LICENCA": "AFASTADO",
        "DISPONIVEL": "DISPONIVEL", "LIVRE": "DISPONIVEL",
        "EM_USO": "EM_USO", "RODANDO": "EM_USO", "OCUPADO": "EM_USO",
        "MANUTENCAO": "MANUTENCAO", "OFICINA": "MANUTENCAO", "QUEBRADO": "MANUTENCAO",
        "EM_ESTOQUE": "EM_ESTOQUE", "DISPONIVEL_ESTOQUE": "EM_ESTOQUE",
        "BAIXO_ESTOQUE": "BAIXO_ESTOQUE", "ACABANDO": "BAIXO_ESTOQUE",
        "ESGOTADO": "ESGOTADO", "SEM_ESTOQUE": "ESGOTADO",
        "RESERVADO": "RESERVADO", "SEPARADO": "RESERVADO",
        "PENDENTE": "PENDENTE", "AGUARDANDO": "PENDENTE",
        "EM_ROTA": "EM_ROTA", "TRANSITO": "EM_ROTA", "CAMINHO": "EM_ROTA",
        "ENTREGUE": "ENTREGUE", "CONCLUIDA_ENTREGA": "ENTREGUE",
        "ATRASADA": "ATRASADA", "ATRASO": "ATRASADA",
        "PRESENTE": "PRESENTE", "RECEBIDO": "RECEBIDO", "CONFERIDO": "CONFERIDO"
    }

    status_permitidos = regras.get(mod, [])
    if not status_permitidos:
        status_encontrado = mapa_geral.get(s_lookup, s_lookup)
        return {"ok": True, "valor": status_encontrado, "motivo": "Status aceito por normalização contextual (Módulo genérico)", "original": original}

    if s_lookup in status_permitidos:
        return {"ok": True, "valor": s_lookup, "motivo": None, "original": original}

    mapa_modulo = mapas_contextuais.get(mod, {})
    if s_lookup in mapa_modulo:
        val_contextual = mapa_modulo[s_lookup]
        if val_contextual in status_permitidos:
            return {"ok": True, "valor": val_contextual, "motivo": "Status normalizado por regra de módulo", "original": original}

    if s_lookup in mapa_geral:
        val_global = mapa_geral[s_lookup]
        if val_global in status_permitidos:
            msg = f"Status normalizado para o padrão do módulo" if s_lookup != val_global else None
            return {"ok": True, "valor": val_global, "motivo": msg, "original": original}

    status_encontrado = mapa_geral.get(s_lookup, s_lookup)
    if status_encontrado in mapa_geral.values():
        return {"ok": True, "valor": status_encontrado, "motivo": "Status aceito por normalização contextual (Tolerância global)", "original": original}
    if len(s_lookup) > 1:
        return {"ok": True, "valor": status_encontrado, "motivo": "Status mantido por tolerância (Fora do padrão conhecido)", "original": original}

    msg_nome_modulo = mod.capitalize() if mod != "OBRAS" else "Obras"
    return {"ok": False, "valor": "INVÁLIDO", "motivo": f"Status vazio ou ilegível não reconhecido no módulo {msg_nome_modulo}", "original": original}


LIMITE_ORDENACAO_LINHAS = 15000
LIMITE_MODO_RAPIDO_LINHAS = 3000
LIMITE_MODO_RAPIDO_CELULAS = 120000
LIMITE_AMOSTRA_LARGURA = 400
LIMITE_REGISTROS_INCONSISTENCIA = 500
LIMITE_LOG_CORRECOES = 3000

BASE_COLUNAS_VALIDACAO = {
    "nome_completo": {"aliases": ["nome", "nome_completo", "cliente", "funcionario", "colaborador"], "tipo": "texto_nome", "required": False, "regras": {"min_len": 2, "max_len": 150}},
    "cpf": {"aliases": ["cpf", "num_cpf", "documento"], "tipo": "cpf", "required": False, "regras": {"formatar_saida": True, "validar_digitos": True}},
    "cnpj": {"aliases": ["cnpj", "num_cnpj", "documento_empresa"], "tipo": "cnpj", "required": False, "regras": {"formatar_saida": True, "validar_digitos": True}},
    "data_nascimento": {"aliases": ["data_nascimento", "nascimento"], "tipo": "data", "required": False, "regras": {"nao_permitir_futuro": True, "ano_minimo": 1900}},
    "idade": {"aliases": ["idade", "anos"], "tipo": "inteiro", "required": False, "regras": {"min": 0, "max": 120}},
    "telefone": {"aliases": ["telefone", "fone", "tel"], "tipo": "telefone_br", "required": False, "regras": {}},
    "celular": {"aliases": ["celular", "cel", "mobile"], "tipo": "telefone_br", "required": False, "regras": {}},
    "whatsapp": {"aliases": ["whatsapp", "whats", "zap"], "tipo": "telefone_br", "required": False, "regras": {}},
    "email": {"aliases": ["email", "e-mail", "mail"], "tipo": "email", "required": False, "regras": {}},
    "cep": {"aliases": ["cep", "codigo_postal"], "tipo": "cep_br", "required": False, "regras": {}},
    "endereco": {"aliases": ["endereco", "endereço", "logradouro", "rua"], "tipo": "texto", "required": False, "regras": {"min_len": 2, "max_len": 255}},
    "bairro": {"aliases": ["bairro", "distrito"], "tipo": "texto", "required": False, "regras": {"min_len": 2, "max_len": 120}},
    "cidade": {"aliases": ["cidade", "municipio", "município"], "tipo": "texto", "required": False, "regras": {"min_len": 2, "max_len": 120}},
    "estado": {"aliases": ["estado", "uf", "sigla_uf"], "tipo": "uf_br", "required": False, "regras": {}},
    "sexo": {"aliases": ["sexo", "genero", "gênero"], "tipo": "lista", "required": False, "regras": {"permitidos": ["M", "F", "OUTRO", "NAO_INFORMADO"]}},
    "motorista": {"aliases": ["motorista", "condutor", "mot"], "tipo": "texto_nome", "required": False, "regras": {"min_len": 2, "max_len": 150}},
    "material": {"aliases": ["material", "nome_material", "produto", "item", "descrição"], "tipo": "texto", "required": False, "regras": {"min_len": 2, "max_len": 255}},
    "unidade": {"aliases": ["unidade", "und", "unid", "medida", "ud"], "tipo": "texto", "required": False, "regras": {"min_len": 1, "max_len": 20}},
    "hora_chegada": {"aliases": ["hora_chegada", "chegada", "entrada", "h_chegada", "hora_entrada"], "tipo": "hora", "required": False, "regras": {}},
    "hora_saida": {"aliases": ["hora_saida", "saida", "saída", "h_saida", "hora_termino", "termino"], "tipo": "hora", "required": False, "regras": {}},
    "hora_prevista": {"aliases": ["hora_prevista", "prevista", "h_prevista", "entrada_prevista"], "tipo": "hora", "required": False, "regras": {}},
    "falta": {"aliases": ["falta", "faltou", "ausencia"], "tipo": "texto", "required": False, "regras": {}},
    "atrasado": {"aliases": ["atrasado", "atraso"], "tipo": "texto", "required": False, "regras": {}},
    "pontual": {"aliases": ["pontual", "no_horario", "na_hora"], "tipo": "texto", "required": False, "regras": {}},
    "quantidade": {"aliases": ["quantidade", "qtd", "qtde", "quant", "volume", "unidades"], "tipo": "quantidade_inteligente", "required": False, "regras": {"min": 0, "max": 500}},
    "estoque": {"aliases": ["estoque", "saldo_estoque", "saldo"], "tipo": "quantidade_inteligente", "required": False, "regras": {"min": 0, "max": 500000}},
    "placa_veiculo": {"aliases": ["placa", "placa_veiculo", "placa_carro"], "tipo": "placa_veiculo", "required": False, "regras": {}},
    "valor_unitario": {"aliases": ["valor_unitario", "preço_unitário", "unitario"], "tipo": "moeda", "required": False, "regras": {"min": 0, "max": 999999999.99}},
    "valor_total": {"aliases": ["valor_total", "total", "montante"], "tipo": "moeda", "required": False, "regras": {"min": 0, "max": 999999999.99}},
    "status": {"aliases": ["status", "situacao", "situação", "st"], "tipo": "lista", "required": False, "regras": {}},
    "data_cadastro": {"aliases": ["data_cadastro", "dt_cadastro", "data"], "tipo": "data", "required": False, "regras": {"ano_minimo": 1900}},
}

TIPOS_VALIDACAO = {
    "texto": {"descricao": "Texto genérico limpo"},
    "texto_nome": {"descricao": "Nome de pessoa/entidade"},
    "inteiro": {"descricao": "Número inteiro"},
    "quantidade_inteligente": {"descricao": "Quantidade processada inteligente"},
    "placa_veiculo": {"descricao": "Placa Automotiva"},
    "decimal": {"descricao": "Número decimal"},
    "moeda": {"descricao": "Valor monetário"},
    "cpf": {"descricao": "CPF brasileiro"},
    "cnpj": {"descricao": "CNPJ brasileiro"},
    "uf_br": {"descricao": "UF brasileira"},
    "cep_br": {"descricao": "CEP brasileiro"},
    "telefone_br": {"descricao": "Telefone brasileiro"},
    "email": {"descricao": "E-mail válido"},
    "data": {"descricao": "Data válida"},
    "data_hora": {"descricao": "Data e hora válidas"},
    "hora": {"descricao": "Hora válida"},
    "lista": {"descricao": "Lista fechada"}
}

UF_VALIDAS = {
    "AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO",
    "MA", "MT", "MS", "MG", "PA", "PB", "PR", "PE", "PI",
    "RJ", "RN", "RS", "RO", "RR", "SC", "SP", "SE", "TO"
}

MAPA_ESTADOS_POR_NOME = {
    "acre": "AC", "alagoas": "AL", "amapa": "AP", "amazonas": "AM", "bahia": "BA",
    "ceara": "CE", "distrito federal": "DF", "espirito santo": "ES", "goias": "GO",
    "maranhao": "MA", "mato grosso": "MT", "mato grosso do sul": "MS", "minas gerais": "MG",
    "para": "PA", "paraiba": "PB", "parana": "PR", "pernambuco": "PE", "piaui": "PI",
    "rio de janeiro": "RJ", "rio grande do norte": "RN", "rio grande do sul": "RS",
    "rondonia": "RO", "roraima": "RR", "santa catarina": "SC", "sao paulo": "SP",
    "sergipe": "SE", "tocantins": "TO"
}


def construir_mapa_aliases(base):
    mapa = {}
    for canonico, meta in base.items():
        aliases = set(meta.get("aliases", [])) | {canonico}
        for alias in aliases:
            mapa[normalizar_nome_coluna(alias)] = canonico
    return mapa


MAPA_ALIAS_CANONICO = construir_mapa_aliases(BASE_COLUNAS_VALIDACAO)


def resolver_coluna_canonica(nome_coluna, modulo_ativo="DESCONHECIDO"):
    norm = normalizar_nome_coluna(nome_coluna)
    canonico = MAPA_ALIAS_CANONICO.get(norm)
    if canonico == "material" and modulo_ativo != "MATERIAIS":
        if "material" not in norm and "produto" not in norm:
            return None
    if canonico == "motorista" and modulo_ativo != "VEICULOS":
        if "motorista" not in norm and "condutor" not in norm:
            return None
    return canonico


def obter_schema_coluna(canonico):
    return BASE_COLUNAS_VALIDACAO.get(canonico, {})


def validar_limite_por_schema(schema, numero):
    regras = schema.get("regras", {})
    minimo = regras.get("min")
    maximo = regras.get("max")
    if minimo is not None and numero < minimo: return False
    if maximo is not None and numero > maximo: return False
    return True


def cpf_valido(digitos):
    if len(digitos) != 11 or digitos == digitos[0] * 11: return False
    soma = sum(int(digitos[i]) * (10 - i) for i in range(9))
    d1 = ((soma * 10) % 11) % 10
    soma = sum(int(digitos[i]) * (11 - i) for i in range(10))
    d2 = ((soma * 10) % 11) % 10
    return digitos[-2:] == f"{d1}{d2}"


def cnpj_valido(digitos):
    if len(digitos) != 14 or digitos == digitos[0] * 14: return False
    def calc(base, pesos):
        soma = sum(int(n) * p for n, p in zip(base, pesos))
        resto = soma % 11
        return "0" if resto < 2 else str(11 - resto)
    d1 = calc(digitos[:12], [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2])
    d2 = calc(digitos[:12] + d1, [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2])
    return digitos[-2:] == d1 + d2


def detectar_estrutura(ws):
    cache = getattr(ws, "_cache_dimensoes_reais", None)
    if cache: return cache

    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    first_row, last_row = None, None

    for r in range(1, max_row + 1):
        if not linha_vazia([ws.cell(r, c).value for c in range(1, max_col + 1)]):
            first_row = r; break

    if first_row is None:
        info = {"header_row": 1, "data_start": 2, "data_end": 1, "col_start": 1, "col_end": 1, "numeric_cols": set()}
        ws._cache_dimensoes_reais = info
        return info

    for r in range(max_row, 0, -1):
        if not linha_vazia([ws.cell(r, c).value for c in range(1, max_col + 1)]):
            last_row = r; break

    first_col, last_col = None, None
    for c in range(1, max_col + 1):
        if any(not valor_vazio(ws.cell(r, c).value) for r in range(first_row, last_row + 1)):
            first_col = c; break

    for c in range(max_col, 0, -1):
        if any(not valor_vazio(ws.cell(r, c).value) for r in range(first_row, last_row + 1)):
            last_col = c; break

    melhor_row, melhor_score = first_row, -1
    for r in range(first_row, min(last_row, first_row + 8) + 1):
        valores = [ws.cell(r, c).value for c in range(first_col, last_col + 1)]
        nao_vazios = [v for v in valores if not valor_vazio(v)]
        if not nao_vazios: continue
        score = sum(3 if isinstance(v, str) else -2 for v in nao_vazios)
        score += len(set(normalizar_nome_coluna(str(v)) for v in nao_vazios))
        if score > melhor_score:
            melhor_score, melhor_row = score, r

    header_row = melhor_row
    data_start = header_row + 1
    data_end = last_row
    col_start = first_col
    col_end = last_col
    numeric_cols = set()

    if data_start <= data_end:
        for c in range(col_start, col_end + 1):
            if any(k in str(ws.cell(header_row, c).value or "").lower() for k in ["cpf", "cnpj", "cep", "telefone", "placa", "id"]): continue
            tot, nums = 0, 0
            for r in range(data_start, min(data_end, data_start + 200) + 1):
                val = ws.cell(r, c).value
                if valor_vazio(val): continue
                tot += 1
                if isinstance(val, (int, float, Decimal)) or parse_decimal_br(val) is not None: nums += 1
            if tot > 0 and (nums / tot) >= 0.7: numeric_cols.add(c)

    info = {"header_row": header_row, "data_start": data_start, "data_end": data_end,
            "col_start": col_start, "col_end": col_end, "numeric_cols": numeric_cols}
    ws._cache_dimensoes_reais = info
    return info


def auto_ajustar_largura(ws):
    info = detectar_estrutura(ws)
    hr, ds, de, cs, ce = info.get("header_row"), info.get("data_start"), info.get("data_end"), info.get("col_start"), info.get("col_end")
    if not hr or cs is None or ce is None or cs > ce: return

    for col in range(cs, ce + 1):
        header_val = str(ws.cell(row=hr, column=col).value or "").strip()
        canonico = resolver_coluna_canonica(header_val)
        tipo = obter_schema_coluna(canonico).get("tipo") if canonico else None
        max_len = len(header_val)

        for row in range(ds, min(de, ds + LIMITE_AMOSTRA_LARGURA) + 1):
            val = str(ws.cell(row=row, column=col).value or "")
            maior_linha = max((len(l.strip()) for l in val.split("\n")), default=0)
            if maior_linha > max_len: max_len = maior_linha

        # Padding premium: mais espaço interno para respirar
        largura_base = max_len + 10
        largura_ideal = max(18, min(largura_base, 56))

        if tipo == "hora":            largura_ideal = max(largura_ideal, 18)
        elif tipo == "data":          largura_ideal = max(largura_ideal, 20)
        elif tipo in ["moeda", "decimal"]: largura_ideal = max(largura_ideal, 22)
        elif tipo == "texto_nome":    largura_ideal = max(largura_ideal, 28)
        elif tipo in ["cpf", "cnpj", "telefone_br"]: largura_ideal = max(largura_ideal, 22)
        elif tipo == "email":         largura_ideal = max(largura_ideal, 32)

        ws.column_dimensions[get_column_letter(col)].width = largura_ideal


def filtrar_colunas_util(ws):
    info = detectar_estrutura(ws)
    hr, cs, ce = info.get("header_row"), info.get("col_start"), info.get("col_end")
    if not hr or cs is None or ce is None or cs > ce: return
    termos_lixo = ["id", "created", "updated", "sample", "uuid", "meta", "system", "hash"]
    exatas_lixo = ["data_criacao", "is_sample", "created_by_id"]
    colunas_removidas = False

    for col in range(ce, cs - 1, -1):
        nome_norm = normalizar_nome_coluna(str(ws.cell(row=hr, column=col).value or ""))
        if nome_norm in exatas_lixo or any(termo in nome_norm.split('_') for termo in termos_lixo):
            ws.delete_cols(col)
            colunas_removidas = True
    if colunas_removidas and hasattr(ws, "_cache_dimensoes_reais"):
        delattr(ws, "_cache_dimensoes_reais")


def obter_ou_criar_aba_inconsistencias(wb):
    nome = "⚠ Inconsistências"
    if nome in wb.sheetnames:
        ws = wb[nome]
    else:
        ws = wb.create_sheet(nome)
        cabecalhos = ["Módulo (Aba)", "Linha", "Coluna Mapeada", "Valor Original", "Valor Atualizado", "Status Final", "Motivo da Rejeição/Correção"]
        for i, h in enumerate(cabecalhos, start=1):
            ws.cell(row=1, column=i, value=h)

    if not hasattr(ws, "_total_registros"): ws._total_registros = max(0, ws.max_row - 1)
    if not hasattr(ws, "_limite_atingido"): ws._limite_atingido = False
    if not hasattr(ws, "_next_row"): ws._next_row = ws.max_row + 1
    return ws


def registrar_inconsistencia(ws_inc, nome_aba, linha, coluna_nome, valor_original, valor_resultado, status_final, motivo):
    total = getattr(ws_inc, "_total_registros", 0)
    limite = getattr(ws_inc, "_limite_atingido", False)

    if total >= LIMITE_REGISTROS_INCONSISTENCIA:
        if not limite:
            ws_inc._limite_atingido = True
            nova_linha = getattr(ws_inc, "_next_row", ws_inc.max_row + 1)
            aviso = ["SISTEMA", "-", "-", "-", "-", "ALERTA", f"Limite de {LIMITE_REGISTROS_INCONSISTENCIA} excedido. Mais ocorrencias foram tratadas em background."]
            for i, valor in enumerate(aviso, start=1): ws_inc.cell(row=nova_linha, column=i, value=valor)
            ws_inc._next_row = nova_linha + 1
        return

    nova_linha = getattr(ws_inc, "_next_row", ws_inc.max_row + 1)
    motivo_limpo = motivo if motivo else "Não especificado"
    dados = [nome_aba, linha, coluna_nome, valor_original, valor_resultado, status_final, motivo_limpo]
    for i, valor in enumerate(dados, start=1):
        ws_inc.cell(row=nova_linha, column=i, value=valor)

    ws_inc._next_row = nova_linha + 1
    ws_inc._total_registros = total + 1


def validar_e_transformar_valor(canonico, valor, nome_aba=None):
    original = valor
    schema = obter_schema_coluna(canonico)
    tipo = schema.get("tipo")
    regras = schema.get("regras", {})

    if not tipo: return {"ok": True, "valor": valor, "motivo": None, "original": original}

    if valor_vazio(valor):
        if schema.get("required"): return {"ok": False, "valor": "INVÁLIDO", "motivo": "Campo obrigatório vazio", "original": original}
        return {"ok": True, "valor": None, "motivo": None, "original": original}

    if tipo == "texto":
        s = normalizar_texto(valor)
        if not s: return {"ok": False, "valor": "INVÁLIDO", "motivo": "Texto vazio", "original": original}
        if len(s) < regras.get("min_len", 1) or len(s) > regras.get("max_len", 255):
            return {"ok": False, "valor": "INVÁLIDO", "motivo": "Texto fora do tamanho permitido", "original": original}
        return {"ok": True, "valor": s, "motivo": "Espaços corrigidos" if s != original else None, "original": original}

    elif tipo == "texto_nome":
        s = normalizar_texto(valor)
        if not s: return {"ok": False, "valor": "INVÁLIDO", "motivo": "Nome vazio", "original": original}
        if len(s) < regras.get("min_len", 2): return {"ok": False, "valor": "INVÁLIDO", "motivo": "Nome muito curto", "original": original}
        novo = s.title()
        return {"ok": True, "valor": novo, "motivo": "Caixa padronizada" if novo != original else None, "original": original}

    elif tipo == "inteiro":
        num = parse_inteiro(valor)
        if num is None: return {"ok": False, "valor": "INVÁLIDO", "motivo": "Inteiro inválido", "original": original}
        if not validar_limite_por_schema(schema, num): return {"ok": False, "valor": "INVÁLIDO", "motivo": "Valor fora do limite permitido", "original": original}
        return {"ok": True, "valor": num, "motivo": None, "original": original}

    elif tipo == "quantidade_inteligente":
        s = str(valor).strip()
        match = re.search(r'\d+', s)
        if not match: return {"ok": False, "valor": "INVÁLIDO", "motivo": "Não foi possível extrair número", "original": original}
        num = int(match.group())
        if num > 500: return {"ok": False, "valor": "INVÁLIDO", "motivo": "Quantidade acima do limite (Máx 500)", "original": original}
        if not validar_limite_por_schema(schema, num): return {"ok": False, "valor": "INVÁLIDO", "motivo": "Quantidade fora do limite permitido", "original": original}
        msg = f"Quantidade filtrada automaticamente" if str(num) != s else None
        return {"ok": True, "valor": num, "motivo": msg, "original": original}

    elif tipo == "placa_veiculo":
        p_formatada = normalizar_placa(valor)
        msg = "Placa padronizada visualmente" if str(original) != str(p_formatada) else None
        return {"ok": True, "valor": p_formatada, "motivo": msg, "original": original}

    elif tipo in ["decimal", "moeda"]:
        num = parse_decimal_br(valor)
        if num is None: return {"ok": False, "valor": "INVÁLIDO", "motivo": "Valor monetário/decimal inválido", "original": original}
        num_float = float(num)
        if not validar_limite_por_schema(schema, num_float): return {"ok": False, "valor": "INVÁLIDO", "motivo": "Valor fora do limite", "original": original}
        return {"ok": True, "valor": num_float, "motivo": None, "original": original}

    elif tipo == "cpf":
        dig = apenas_digitos(valor)
        if len(dig) != 11: return {"ok": False, "valor": "INVÁLIDO", "motivo": "CPF precisa ter 11 dígitos", "original": original}
        if regras.get("validar_digitos", True) and not cpf_valido(dig): return {"ok": False, "valor": "INVÁLIDO", "motivo": "CPF inválido", "original": original}
        return {"ok": True, "valor": formatar_cpf(dig) if regras.get("formatar_saida", True) else dig, "motivo": None, "original": original}

    elif tipo == "cnpj":
        dig = apenas_digitos(valor)
        if len(dig) != 14: return {"ok": False, "valor": "INVÁLIDO", "motivo": "CNPJ precisa ter 14 dígitos", "original": original}
        if regras.get("validar_digitos", True) and not cnpj_valido(dig): return {"ok": False, "valor": "INVÁLIDO", "motivo": "CNPJ inválido", "original": original}
        return {"ok": True, "valor": formatar_cnpj(dig) if regras.get("formatar_saida", True) else dig, "motivo": None, "original": original}

    elif tipo == "cep_br":
        dig = apenas_digitos(valor)
        if len(dig) != 8: return {"ok": False, "valor": "INVÁLIDO", "motivo": "CEP inválido", "original": original}
        return {"ok": True, "valor": formatar_cep(dig), "motivo": None, "original": original}

    elif tipo == "telefone_br":
        dig = apenas_digitos(valor)
        fmt = formatar_telefone_br(dig)
        if not fmt: return {"ok": False, "valor": "INVÁLIDO", "motivo": "Telefone inválido", "original": original}
        return {"ok": True, "valor": fmt, "motivo": None, "original": original}

    elif tipo == "email":
        s = re.sub(r"[^a-z0-9._\-@]", "", str(valor or "").lower().replace(" ", ""))
        if eh_email_valido(s): return {"ok": True, "valor": s, "motivo": None, "original": original}
        return {"ok": False, "valor": "INVÁLIDO", "motivo": "E-mail inválido", "original": original}

    elif tipo == "uf_br":
        s_lookup = _remover_acentos(normalizar_texto(valor)).lower()
        s = MAPA_ESTADOS_POR_NOME.get(s_lookup, s_lookup.upper())
        if s in UF_VALIDAS: return {"ok": True, "valor": s, "motivo": None, "original": original}
        return {"ok": False, "valor": "INVÁLIDO", "motivo": "UF inválida", "original": original}

    elif tipo in ["data", "data_hora"]:
        dt = _converter_texto_para_data(valor)
        if dt is None: return {"ok": False, "valor": "INVÁLIDO", "motivo": "Data inválida ou formato desconhecido", "original": original}
        if regras.get("nao_permitir_futuro", False) and dt.date() > datetime.now().date():
            return {"ok": False, "valor": "INVÁLIDO", "motivo": "Data futura não permitida", "original": original}
        ano_min = regras.get("ano_minimo")
        if ano_min is not None and dt.year < ano_min:
            return {"ok": False, "valor": "INVÁLIDO", "motivo": "Data muito antiga", "original": original}
        return {"ok": True, "valor": dt, "motivo": "Data padronizada" if isinstance(original, str) else None, "original": original}

    elif tipo == "lista":
        if canonico == "status": return normalizar_status_por_modulo(valor, nome_aba, original)
        return {"ok": True, "valor": valor, "motivo": None, "original": original}

    return {"ok": True, "valor": valor, "motivo": None, "original": original}


def validador_veiculos(linha_ctx):
    if "motorista" in linha_ctx:
        ctx = linha_ctx["motorista"]
        if ctx.get("ok") and str(ctx.get("v_atual")).upper() == "ELE MESMO":
            ctx["v_atual"] = "SEM MOTORISTA"
            ctx["motivo"] = "Motorista convertido para termo padrão inteligente"
            ctx["corrigido"] = True

def validador_materiais(linha_ctx):
    if "material" in linha_ctx and "unidade" in linha_ctx:
        mat = linha_ctx["material"]
        uni = linha_ctx["unidade"]
        if mat.get("ok") and "CIMENTO" in normalizar_texto_upper_sem_acentos(mat.get("v_atual")):
            uni["v_atual"] = "SACO"
            uni["ok"] = True
            uni["motivo"] = "Unidade de cimento definida para SACO"
            uni["corrigido"] = True

def validador_equipe(linha_ctx):
    if "falta" in linha_ctx:
        f_val = normalizar_texto_upper_sem_acentos(linha_ctx["falta"].get("v_atual", ""))
        if f_val in ["SIM", "S", "VERDADEIRO", "X", "1"]:
            linha_ctx["falta"]["v_atual"] = "SIM"
            linha_ctx["falta"]["ok"] = True
            for col in ["atrasado", "pontual", "hora_chegada"]:
                if col in linha_ctx:
                    linha_ctx[col]["v_atual"] = None
                    linha_ctx[col]["motivo"] = "Campo ignorado porque registro está marcado como Falta"
                    linha_ctx[col]["ok"] = True
                    linha_ctx[col]["corrigido"] = True
            return

    if "hora_chegada" in linha_ctx and "hora_prevista" in linha_ctx:
        chegada = linha_ctx["hora_chegada"].get("v_atual")
        prevista = linha_ctx["hora_prevista"].get("v_atual")
        if isinstance(chegada, time) and isinstance(prevista, time):
            houve_atraso = chegada > prevista
            if "atrasado" in linha_ctx:
                linha_ctx["atrasado"]["v_atual"] = "SIM" if houve_atraso else "NÃO"
                linha_ctx["atrasado"]["ok"] = True
                linha_ctx["atrasado"]["corrigido"] = True
                linha_ctx["atrasado"]["motivo"] = "Atraso recalculado com base em Hora Chegada e Hora Prevista"
            if "pontual" in linha_ctx:
                linha_ctx["pontual"]["v_atual"] = "NÃO" if houve_atraso else "SIM"
                linha_ctx["pontual"]["ok"] = True
                linha_ctx["pontual"]["corrigido"] = True
                linha_ctx["pontual"]["motivo"] = "Pontualidade recalculada com base em Hora Chegada e Hora Prevista"


VALIDADORES_MODULO = {
    "VEICULOS": validador_veiculos,
    "MATERIAIS": validador_materiais,
    "OBRAS": lambda ctx: None,
    "EQUIPE": validador_equipe,
    "ENTREGAS": lambda ctx: None
}


def validar_sheet(ws, ws_inc):
    info = detectar_estrutura(ws)
    hr, ds, de, cs, ce = info["header_row"], info["data_start"], info["data_end"], info["col_start"], info["col_end"]
    if not hr or ds > de: return {"total_colunas_mapeadas": 0, "total_invalidos": 0, "total_corrigidos": 0}

    colunas_map = {}
    modulo_ativo = detectar_modulo_por_aba(ws.title)

    for col in range(cs, ce + 1):
        nome_coluna = ws.cell(hr, col).value
        canonico = resolver_coluna_canonica(nome_coluna, modulo_ativo)
        if canonico:
            colunas_map[col] = {"nome_original": str(nome_coluna or ""), "canonico": canonico, "required": obter_schema_coluna(canonico).get("required", False)}

    if not colunas_map: return {"total_colunas_mapeadas": 0, "total_invalidos": 0, "total_corrigidos": 0}

    colunas_validadas = sorted(colunas_map.keys())
    colunas_relativas = [col - cs for col in colunas_validadas]
    total_invalidos, total_corrigidos = 0, 0
    logar_correcao = (de - ds + 1) <= LIMITE_LOG_CORRECOES

    for idx_linha, row_cells in enumerate(ws.iter_rows(min_row=ds, max_row=de, min_col=cs, max_col=ce), start=ds):
        if all(valor_vazio(c.value) for c in row_cells): continue
        if not any(not valor_vazio(row_cells[idx].value) for idx in colunas_relativas): continue

        linha_ctx = {}
        for col, rel_idx in zip(colunas_validadas, colunas_relativas):
            cell, meta = row_cells[rel_idx], colunas_map[col]
            valor_antes = cell.value
            if obter_schema_coluna(meta["canonico"]).get("tipo") != "hora":
                res = validar_e_transformar_valor(meta["canonico"], valor_antes, nome_aba=ws.title)
                linha_ctx[meta["canonico"]] = {
                    "cell": cell, "meta": meta, "v_orig": valor_antes,
                    "v_atual": res.get("valor"), "ok": res.get("ok", True),
                    "motivo": res.get("motivo"),
                    "corrigido": (str(valor_antes) != str(res.get("valor")) and res.get("ok", True))
                }

        if modulo_ativo in VALIDADORES_MODULO:
            VALIDADORES_MODULO[modulo_ativo](linha_ctx)

        for col, rel_idx in zip(colunas_validadas, colunas_relativas):
            cell, meta = row_cells[rel_idx], colunas_map[col]
            v_orig = cell.value
            canonico = meta["canonico"]
            schema = obter_schema_coluna(canonico)
            tipo = schema.get("tipo")

            if tipo == "hora":
                hora = processar_hora(v_orig)
                if hora:
                    cell.value = hora
                    cell.number_format = "hh:mm"
                    if str(v_orig) != str(hora):
                        total_corrigidos += 1
                        if logar_correcao:
                            registrar_inconsistencia(ws_inc, ws.title, idx_linha, meta["nome_original"], v_orig, hora, "CORRIGIDO", "Normalização de Hora Automática")
                else:
                    if not valor_vazio(v_orig) or schema.get("required"):
                        cell.value = None
                        total_invalidos += 1
                        registrar_inconsistencia(ws_inc, ws.title, idx_linha, meta["nome_original"], v_orig, None, "INVÁLIDO", "Hora inválida")
                        cell.comment = Comment("Hora inválida", "Data Studio V8")
                continue

            if canonico in linha_ctx:
                dados = linha_ctx[canonico]
                if valor_vazio(v_orig) and not meta["required"]: continue

                if not dados.get("ok", True):
                    total_invalidos += 1
                    if _set_cell_if_changed(cell, "INVÁLIDO") or v_orig == "INVÁLIDO":
                        msg_erro = dados.get("motivo") or "Fora do padrão"
                        cell.comment = Comment(msg_erro, "Data Studio V8")
                        registrar_inconsistencia(ws_inc, ws.title, idx_linha, meta["nome_original"], v_orig, "INVÁLIDO", "INVÁLIDO", msg_erro)
                else:
                    if _set_cell_if_changed(cell, dados["v_atual"]):
                        total_corrigidos += 1
                        if logar_correcao and dados.get("corrigido"):
                            registrar_inconsistencia(ws_inc, ws.title, idx_linha, meta["nome_original"], v_orig, dados["v_atual"], "CORRIGIDO", dados.get("motivo") or "Normalização de Formato Automática")

    return {"total_colunas_mapeadas": len(colunas_map), "total_invalidos": total_invalidos, "total_corrigidos": total_corrigidos}


def ordenar_por_data_na_planilha(ws, hr, ds, de, cs, ce):
    if not hr or ds > de or (de - ds + 1) > LIMITE_ORDENACAO_LINHAS: return

    modulo_ativo = detectar_modulo_por_aba(ws.title)
    col_data = None

    for col in range(cs, ce + 1):
        canonico = resolver_coluna_canonica(ws.cell(hr, col).value, modulo_ativo)
        if canonico and obter_schema_coluna(canonico).get("tipo") in ["data", "data_hora"]:
            col_data = col
            break

    if not col_data: return
    linhas, encontrou_data = [], False
    for row in range(ds, de + 1):
        valores = [ws.cell(row, c).value for c in range(cs, ce + 1)]
        dv = _converter_texto_para_data(valores[col_data - cs])
        if isinstance(dv, datetime) and dv.year <= 1905: dv = None
        if dv is not None: encontrou_data = True
        linhas.append((dv, valores))
    if not encontrou_data: return
    linhas.sort(key=lambda x: x[0] or datetime.max)
    for i, (_, valores) in enumerate(linhas, start=ds):
        for j, valor in enumerate(valores, start=cs):
            ws.cell(row=i, column=j, value=valor if valor is not None else None)


# ─────────────────────────────────────────────────────────────────────────────
#  SISTEMA DE CORES DE STATUS — REDESENHADO
#
#  Filosofia: status não é só cor. É comunicação.
#  Cada cor carrega significado universal:
#    Verde  → sucesso, ativo, positivo
#    Azul   → processo, neutro-informativo
#    Âmbar  → atenção, alerta moderado
#    Vermelho → problema, falha, bloqueio
#    Roxo   → especial, programado
#    Cinza  → inativo, neutro
#
#  Paleta ajustada para elegância: fundos pastel suaves + texto bem escuro.
#  Nada saturado. Nada gritante. Identidade consistente.
# ─────────────────────────────────────────────────────────────────────────────

_PALETA_STATUS = {
    # ── VERDE: positivo / ativo / ok ─────────────────────────────────────
    "DISPONIVEL":    {"bg": "DCFCE7", "fg": "14532D"},
    "CONCLUIDA":     {"bg": "DCFCE7", "fg": "14532D"},
    "ATIVO":         {"bg": "DCFCE7", "fg": "14532D"},
    "PRESENTE":      {"bg": "DCFCE7", "fg": "14532D"},
    "EM_ESTOQUE":    {"bg": "DCFCE7", "fg": "14532D"},
    "ENTREGUE":      {"bg": "DCFCE7", "fg": "14532D"},
    "RECEBIDO":      {"bg": "DCFCE7", "fg": "14532D"},
    "CONFERIDO":     {"bg": "DCFCE7", "fg": "14532D"},

    # ── AZUL ROYAL: processo / em andamento ─────────────────────────────
    "EM_ANDAMENTO":  {"bg": "DBEAFE", "fg": "1E3A8A"},
    "EM_USO":        {"bg": "DBEAFE", "fg": "1E3A8A"},
    "EM_ROTA":       {"bg": "DBEAFE", "fg": "1E3A8A"},
    "RESERVADO":     {"bg": "DBEAFE", "fg": "1E3A8A"},

    # ── ÂMBAR: atenção / pendente ────────────────────────────────────────
    "PARALISADA":    {"bg": "FEF3C7", "fg": "78350F"},
    "MANUTENCAO":    {"bg": "FEF3C7", "fg": "78350F"},
    "AFASTADO":      {"bg": "FEF3C7", "fg": "78350F"},
    "PENDENTE":      {"bg": "FEF3C7", "fg": "78350F"},
    "BAIXO_ESTOQUE": {"bg": "FEF3C7", "fg": "78350F"},

    # ── VERMELHO: problema / falha / crítico ─────────────────────────────
    "CANCELADA":     {"bg": "FEE2E2", "fg": "7F1D1D"},
    "FALTA":         {"bg": "FEE2E2", "fg": "7F1D1D"},
    "ESGOTADO":      {"bg": "FEE2E2", "fg": "7F1D1D"},
    "ATRASADA":      {"bg": "FEE2E2", "fg": "7F1D1D"},

    # ── ROXO: especial / programado ──────────────────────────────────────
    "FERIAS":        {"bg": "EDE9FE", "fg": "4C1D95"},

    # ── CINZA: inativo / neutro ──────────────────────────────────────────
    "INATIVO":       {"bg": "F3F4F6", "fg": "374151"},
}

_CACHE_FILLS_STATUS = {
    k: (
        PatternFill("solid", fgColor=v["bg"]),
        Font(name="Calibri", bold=True, size=10, color=v["fg"])
    )
    for k, v in _PALETA_STATUS.items()
}

# ── ESTILO DE ERRO ────────────────────────────────────────────────────────────
# Vermelho intenso com fundo róseo muito suave. Legível mas alarmante.
_FILL_ERRO  = PatternFill("solid", fgColor="FEE2E2")
_FONT_ERRO  = Font(name="Calibri", bold=True, size=10, color="991B1B", italic=True)
_ALIGN_ERRO = Alignment(horizontal="center", vertical="center")

# ── ESTILOS DE MOEDA ─────────────────────────────────────────────────────────
# Fundos muito sutis para diferenciar sem poluir
_FILL_MOEDA_PAR   = PatternFill("solid", fgColor="F8FFFE")
_FILL_MOEDA_IMPAR = PatternFill("solid", fgColor="FFFFFF")


def _obter_cor_status(val_status_str):
    chave = str(val_status_str).strip().upper().replace(" ", "_") if val_status_str else ""
    return _CACHE_FILLS_STATUS.get(chave, (None, None))


def _aplicar_coloracao_celula(cell, valor, row_fill, normal_font, canonico, status_col_idx, tipo, is_linha_par, modo_rapido):
    if valor == "INVÁLIDO":
        cell.fill = _FILL_ERRO
        cell.font = _FONT_ERRO
        cell.alignment = _ALIGN_ERRO
        return _FILL_ERRO
    if canonico == "status":
        fill_s, font_s = _obter_cor_status(valor)
        if fill_s is not None:
            cell.fill = fill_s
            cell.font = font_s
            cell.alignment = Alignment(horizontal="center", vertical="center")
            return fill_s
    if tipo in ["moeda", "decimal"] and not modo_rapido:
        cell.fill = _FILL_MOEDA_PAR if is_linha_par else _FILL_MOEDA_IMPAR
        cell.font = normal_font
        return cell.fill

    cell.fill = row_fill
    cell.font = normal_font
    return row_fill


# ─────────────────────────────────────────────────────────────────────────────
#  formatar_sheet — REDESENHADO COMPLETAMENTE
#
#  Decisões de design:
#
#  1. CABEÇALHO
#     • Fundo cor âncora do tema (quase-preto ou deep color)
#     • Texto branco / off-white com peso bold size 11
#     • Altura: 42px — presença nobre sem ser exagerado
#     • Alinhamento central / wrapping desabilitado (texto clean)
#     • Borda inferior "medium" na cor tema = linha de separação premium
#
#  2. DADOS
#     • Zebra sutil: linha par = tom muito levemente colorido, ímpar = branco
#     • Linha height padrão: 24px (mais confortável que o padrão Excel)
#     • Indent 1 em texto, centralizado em códigos/datas/status
#     • Fonte Calibri 10 — universal, clean, legível
#     • Nenhuma borda vertical em dados (minimalismo)
#     • Borda horizontal muito sutil nas linhas
#
#  3. BORDAS
#     • Cabeçalho: all sides thin na cor tema
#     • Dados: só borda inferior thin levíssima (cor border do tema)
#     • Última linha: borda inferior medium na cor âncora do tema
#     • Zero borda lateral em dados = visual mais clean, menos grade
#
#  4. COLUNAS DE STATUS
#     • Tag colorida centralizada (sem borda lateral = pill-like)
#
#  5. NÚMEROS / MOEDA
#     • Alinhamento à direita
#     • Formato #,##0.00 (padrão internacional pt-BR)
#
# ─────────────────────────────────────────────────────────────────────────────

def formatar_sheet(ws, tema_nome, ordenar=True):
    t = TEMAS.get(tema_nome, TEMAS["🔵 Azul Executivo"])
    info = detectar_estrutura(ws)
    hr = info["header_row"]
    ds = info["data_start"]
    de = info["data_end"]
    cs = info["col_start"]
    ce = info["col_end"]
    nc = info["numeric_cols"]

    total_linhas   = max(0, de - ds + 1)
    total_colunas  = max(0, ce - cs + 1)
    modo_rapido    = (total_linhas >= LIMITE_MODO_RAPIDO_LINHAS or
                      total_linhas * total_colunas >= LIMITE_MODO_RAPIDO_CELULAS)

    if ordenar and not modo_rapido:
        ordenar_por_data_na_planilha(ws, hr, ds, de, cs, ce)

    # ── Configurações gerais da aba ──────────────────────────────────────────
    ws.sheet_view.showGridLines  = False   # sem grade padrão do Excel
    ws.sheet_view.zoomScale      = 100
    ws.sheet_properties.tabColor = t["tab_color"]

    # ── Fills ────────────────────────────────────────────────────────────────
    header_fill  = _fill(t["header_bg"])
    zebra_fill   = _fill(t["zebra"])
    white_fill   = _fill(t["row_white"])
    panel_fill   = _fill(t["panel_bg"])

    # ── Bordas ───────────────────────────────────────────────────────────────
    # Dados: só borda inferior — clean, respira melhor
    borda_dados_linha = Border(
        bottom=Side(border_style="thin", color=t["border"]),
        left=Side(border_style="thin", color=t["border"]),
        right=Side(border_style="thin", color=t["border"]),
    )
    # Cabeçalho: borda completa fina + borda inferior forte
    borda_header = Border(
        left  =Side(border_style="thin",   color=t["header_bg"]),
        right =Side(border_style="thin",   color=t["header_bg"]),
        top   =Side(border_style="thin",   color=t["header_bg"]),
        bottom=Side(border_style="medium", color=t["header_bg"]),
    )
    # Última linha da tabela: borda inferior forte (fecho visual)
    borda_ultima_linha = Border(
        left  =Side(border_style="thin",   color=t["border"]),
        right =Side(border_style="thin",   color=t["border"]),
        top   =Side(border_style="thin",   color=t["border"]),
        bottom=Side(border_style="medium", color=t["border_strong"]),
    )

    # ── Fontes ───────────────────────────────────────────────────────────────
    # Cabeçalho: Calibri Bold 11, off-white — presença sem grito
    header_font = Font(name="Calibri", bold=True, size=11, color=t["header_fg"])
    # Dados normais: Calibri 10, quase-preto
    normal_font = Font(name="Calibri", size=10, color="1C2B3A")
    # Dados numéricos: leve itálico remove para manter profissionalismo
    num_font    = Font(name="Calibri", size=10, color="1C2B3A")

    modulo_ativo = detectar_modulo_por_aba(ws.title)
    colunas_meta = {}
    status_col_idx = None

    # ── CABEÇALHO ─────────────────────────────────────────────────────────────
    if hr:
        # Altura generosa — cria presença visual forte
        ws.row_dimensions[hr].height = 42

        for col in range(cs, ce + 1):
            cell = ws.cell(hr, col)

            # Normaliza título: Title Case limpo
            if isinstance(cell.value, str):
                cell.value = cell.value.strip().title()

            nome_coluna = str(cell.value or "")
            canonico    = resolver_coluna_canonica(nome_coluna, modulo_ativo)
            schema      = obter_schema_coluna(canonico) if canonico else {}
            tipo        = schema.get("tipo")

            if canonico == "status":
                status_col_idx = col

            colunas_meta[col] = {
                "canonico":   canonico,
                "tipo":       tipo,
                "centralizar": tipo in [
                    "hora", "data", "data_hora", "uf_br",
                    "placa_veiculo", "lista", "cep_br", "telefone_br"
                ] or canonico in ["unidade", "idade"]
            }

            cell.fill      = header_fill
            cell.font      = header_font
            cell.border    = borda_header
            # Cabeçalho: centralizado horizontalmente, centrado verticalmente
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    limite_amostra = min(de, ds + LIMITE_AMOSTRA_LARGURA - 1) if ds <= de else de

    # ── LINHAS DE DADOS ───────────────────────────────────────────────────────
    for row in range(ds, de + 1):
        is_ultima_linha = (row == de)
        is_linha_par    = (row - ds) % 2 == 0
        row_fill        = zebra_fill if is_linha_par else white_fill
        # Altura confortável: 26px padrão, aumenta se conteúdo exige
        altura_calculada = 22 if modo_rapido else 26

        for col in range(cs, ce + 1):
            cell  = ws.cell(row, col)
            valor = cell.value

            meta      = colunas_meta.get(col, {})
            tipo      = meta.get("tipo")
            canonico  = meta.get("canonico")

            # ── Correção datas antigas do Excel (bug 1899) ────────────────
            if (not modo_rapido
                    and isinstance(valor, (datetime, date))
                    and not isinstance(valor, time)
                    and tipo != "hora"):
                if valor.year <= 1905:
                    valor_dt = (datetime.combine(valor, datetime.min.time())
                                if isinstance(valor, date) and not isinstance(valor, datetime)
                                else valor)
                    delta = valor_dt - datetime(1899, 12, 30)
                    num   = delta.days + (delta.seconds / 86400.0)
                    valor = cell.value = int(num) if num == int(num) else num

            # ── Formatação por tipo ───────────────────────────────────────
            if tipo in ["data", "data_hora"]:
                d_conv = valor if isinstance(valor, datetime) else _converter_texto_para_data(valor)
                if d_conv is not None and d_conv.year > 1905:
                    cell.value        = d_conv
                    cell.number_format = "DD/MM/YYYY"
                    cell.alignment    = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = _align_indent("left", "center", wrap=not modo_rapido, indent=1)

            elif tipo == "hora" or isinstance(valor, time):
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(valor, time):
                    cell.number_format = "hh:mm"

            elif tipo in ["moeda", "decimal"] and isinstance(valor, (int, float, Decimal)):
                cell.alignment    = Alignment(horizontal="right", vertical="center", indent=1)
                cell.number_format = "R$ #,##0.00"  # formato BR explícito

            elif col in nc and isinstance(valor, (int, float, Decimal)):
                cell.alignment    = Alignment(horizontal="right", vertical="center", indent=1)
                cell.number_format = ("#,##0.00" if isinstance(valor, float) and valor != int(valor)
                                       else "#,##0")

            else:
                if meta.get("centralizar"):
                    cell.alignment = Alignment(horizontal="center", vertical="center",
                                               wrap_text=not modo_rapido)
                else:
                    cell.alignment = _align_indent("left", "center",
                                                   wrap=not modo_rapido, indent=1)

            # ── Borda ─────────────────────────────────────────────────────
            cell.border = borda_ultima_linha if is_ultima_linha else borda_dados_linha

            # ── Cor / Fill ────────────────────────────────────────────────
            _aplicar_coloracao_celula(
                cell, valor, row_fill, normal_font,
                canonico, status_col_idx, tipo, is_linha_par, modo_rapido
            )

            # ── Altura dinâmica por conteúdo ──────────────────────────────
            if not modo_rapido and row <= limite_amostra:
                LARGURA_BASE_TELA = 44
                tamanho_texto = len(str(cell.value)) if cell.value is not None else 0
                if tamanho_texto > LARGURA_BASE_TELA:
                    linhas_totais = max(
                        (tamanho_texto // LARGURA_BASE_TELA) + 1,
                        str(cell.value).count('\n') + 1
                    )
                    altura_com_respiro = (linhas_totais * 16) + 16
                    if altura_com_respiro > altura_calculada:
                        altura_calculada = altura_com_respiro

        ws.row_dimensions[row].height = altura_calculada

    # ── Freeze + Filtro ──────────────────────────────────────────────────────
    if hr:
        ws.freeze_panes = ws.cell(hr + 1, cs)
        if ds <= de and cs <= ce:
            ws.auto_filter.ref = (
                f"{get_column_letter(cs)}{hr}:{get_column_letter(ce)}{de}"
            )


# ─────────────────────────────────────────────────────────────────────────────
#  criar_sumario — DASHBOARD PREMIUM COMPLETO
#
#  Filosofia: o Dashboard não é uma lista. É um painel executivo.
#  Layout de cards com métricas, separação visual clara, identidade forte.
#
#  Estrutura:
#    • Linha 1-2: área em branco (respiro superior)
#    • Linha 3: título principal + subtítulo
#    • Linha 4: separador visual
#    • Linha 5-7: cards de métricas (Total Abas / Total Linhas / Data)
#    • Linha 9: cabeçalho da tabela de abas
#    • Linha 10+: listagem de abas formatada
# ─────────────────────────────────────────────────────────────────────────────

def criar_sumario(wb, tema_nome):
    t = TEMAS.get(tema_nome, TEMAS["🔵 Azul Executivo"])

    if "📊 Dashboard" in wb.sheetnames:
        del wb["📊 Dashboard"]

    ws = wb.create_sheet("📊 Dashboard", 0)
    ws.sheet_view.showGridLines  = False
    ws.sheet_properties.tabColor = t["tab_color"]

    # ── Larguras das colunas ─────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 3    # margem esquerda
    ws.column_dimensions["B"].width = 6    # índice
    ws.column_dimensions["C"].width = 36   # nome da aba
    ws.column_dimensions["D"].width = 18   # linhas
    ws.column_dimensions["E"].width = 18   # colunas
    ws.column_dimensions["F"].width = 3    # margem direita

    # ── Alturas ──────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height  = 16   # respiro topo
    ws.row_dimensions[2].height  = 14
    ws.row_dimensions[3].height  = 52   # linha do título
    ws.row_dimensions[4].height  = 8    # separador
    ws.row_dimensions[5].height  = 14
    ws.row_dimensions[6].height  = 54   # cards de métrica
    ws.row_dimensions[7].height  = 12
    ws.row_dimensions[8].height  = 16
    ws.row_dimensions[9].height  = 36   # cabeçalho da tabela

    # ── Fills base ───────────────────────────────────────────────────────────
    fill_fundo        = _fill("FFFFFF")
    fill_header_bg    = _fill(t["header_bg"])
    fill_panel        = _fill(t["panel_bg"])
    fill_card         = _fill("FFFFFF")
    fill_zebra        = _fill(t["zebra"])

    # Pinta fundo geral (linhas 1-9) com cor de painel
    for r in range(1, 10):
        for c in range(1, 7):
            ws.cell(r, c).fill = fill_panel

    # ── TÍTULO ────────────────────────────────────────────────────────────────
    ws.merge_cells("B3:E3")
    titulo_cell = ws["B3"]
    titulo_cell.value     = "DASHBOARD DO ARQUIVO"
    titulo_cell.font      = Font(name="Calibri", bold=True, size=22, color=t["title_fg"])
    titulo_cell.alignment = Alignment(horizontal="left", vertical="center")
    titulo_cell.fill      = fill_panel

    # Linha decorativa sob o título (simula underline premium)
    for c in range(2, 6):
        cell = ws.cell(4, c)
        cell.fill   = _fill(t["header_bg"])
        cell.border = Border()

    # ── CARDS DE MÉTRICAS ─────────────────────────────────────────────────────
    abas_processaveis = [s for s in wb.sheetnames if s not in ("📊 Dashboard", "⚠ Inconsistências")]
    total_linhas_geral = sum(
        max(0, detectar_estrutura(wb[a])["data_end"] - detectar_estrutura(wb[a])["data_start"] + 1)
        for a in abas_processaveis
    )

    metricas = [
        ("C6", "TOTAL DE ABAS",    str(len(abas_processaveis))),
        ("D6", "REGISTROS",        f"{total_linhas_geral:,}".replace(",", ".")),
        ("E6", "GERADO EM",        datetime.now().strftime("%d/%m/%Y")),
    ]

    for addr, label, valor in metricas:
        col_letter = addr[0]
        row_num    = int(addr[1:])
        c = ws[addr]
        c.fill      = fill_card
        c.alignment = Alignment(horizontal="center", vertical="center")

        # Label pequeno acima
        label_row = row_num - 0  # dentro da célula mesclamos ou usamos 2 linhas via formato
        # Vamos usar uma célula com \n (quebra)
        c.value = f"{label}\n{valor}"
        c.font  = Font(name="Calibri", bold=False, size=9, color=t["label_fg"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Borda card premium
        borda_card = Border(
            left  =Side(border_style="medium", color=t["card_border"]),
            right =Side(border_style="medium", color=t["card_border"]),
            top   =Side(border_style="medium", color=t["card_border"]),
            bottom=Side(border_style="medium", color=t["card_border"]),
        )
        c.border = borda_card

        # Sobrescreve com formatação do número (valor maior)
        # Usamos duas células: uma para label, uma para valor
        # Estratégia: a célula principal mostra o VALOR grande
        c.value = valor
        c.font  = Font(name="Calibri", bold=True, size=20, color=t["metric_fg"])
        c.alignment = Alignment(horizontal="center", vertical="center")

        # Célula de label (linha acima — row 5)
        label_cell = ws.cell(5, ord(col_letter) - ord("A") + 1)
        label_cell.value     = label
        label_cell.font      = Font(name="Calibri", bold=False, size=9,
                                    color=t["label_fg"], italic=True)
        label_cell.alignment = Alignment(horizontal="center", vertical="bottom")
        label_cell.fill      = fill_panel

    # ── CABEÇALHO DA TABELA ───────────────────────────────────────────────────
    cabecalhos_tabela = ["", "#", "Nome da Planilha", "Registros", "Colunas"]
    for i, h in enumerate(cabecalhos_tabela, start=1):
        cell = ws.cell(9, i)
        cell.value     = h
        cell.fill      = fill_header_bg
        cell.font      = Font(name="Calibri", bold=True, size=10, color=t["header_fg"])
        cell.alignment = Alignment(
            horizontal="center" if i not in (3,) else "left",
            vertical="center",
            indent=1 if i == 3 else 0
        )
        cell.border = Border(
            bottom=Side(border_style="medium", color=t["header_bg"])
        )

    # ── LINHAS DE DADOS DA TABELA ─────────────────────────────────────────────
    for idx, nome in enumerate(abas_processaveis, start=1):
        row     = 9 + idx
        aba     = wb[nome]
        info_a  = detectar_estrutura(aba)
        n_linhas = max(0, info_a["data_end"] - info_a["data_start"] + 1)
        n_colunas = max(0, info_a["col_end"] - info_a["col_start"] + 1)

        is_par    = idx % 2 == 0
        row_fill  = fill_zebra if is_par else _fill("FFFFFF")

        dados_row = ["", f"{idx:02d}", nome, f"{n_linhas:,}".replace(",", "."), str(n_colunas)]
        aligns    = ["center", "center", "left", "center", "center"]
        indents   = [0, 0, 1, 0, 0]

        for j, (val, halign, ind) in enumerate(zip(dados_row, aligns, indents), start=1):
            cell = ws.cell(row, j)
            cell.value     = val
            cell.fill      = row_fill
            cell.font      = Font(name="Calibri", size=10, color="1C2B3A")
            cell.alignment = Alignment(horizontal=halign, vertical="center", indent=ind)
            cell.border    = Border(
                bottom=Side(border_style="thin", color=t["border"])
            )

        ws.row_dimensions[row].height = 26

    # Linha de fechamento da tabela
    linha_fechamento = 9 + len(abas_processaveis) + 1
    for c in range(1, 6):
        cell = ws.cell(linha_fechamento, c)
        cell.fill   = fill_panel
        cell.border = Border(top=Side(border_style="medium", color=t["border_strong"]))


def criar_resumo_consolidacao(wb, arquivos_usados, total_linhas, resumo_categorias, tema_nome):
    t = TEMAS.get(tema_nome, TEMAS["🔵 Azul Executivo"])

    if "📊 Visão Geral" in wb.sheetnames:
        del wb["📊 Visão Geral"]

    ws = wb.create_sheet("📊 Visão Geral", 0)
    ws.sheet_view.showGridLines  = False
    ws.sheet_properties.tabColor = t["tab_color"]

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 42
    ws.column_dimensions["E"].width = 3

    fill_panel   = _fill(t["panel_bg"])
    fill_header  = _fill(t["header_bg"])
    fill_zebra   = _fill(t["zebra"])

    for r in range(1, 9):
        for c in range(1, 6):
            ws.cell(r, c).fill = fill_panel

    ws.row_dimensions[1].height = 16
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 52
    ws.row_dimensions[4].height = 8
    ws.row_dimensions[5].height = 14
    ws.row_dimensions[6].height = 54
    ws.row_dimensions[7].height = 16
    ws.row_dimensions[8].height = 36

    ws.merge_cells("B3:D3")
    t_cell = ws["B3"]
    t_cell.value     = "CONSOLIDAÇÃO EXECUTIVA"
    t_cell.font      = Font(name="Calibri", bold=True, size=22, color=t["title_fg"])
    t_cell.alignment = Alignment(horizontal="left", vertical="center")
    t_cell.fill      = fill_panel

    for c in range(2, 5):
        cell = ws.cell(4, c)
        cell.fill = _fill(t["header_bg"])

    metricas_cons = [
        ("B6", "ARQUIVOS",   str(len(arquivos_usados))),
        ("C6", "REGISTROS",  f"{total_linhas:,}".replace(",", ".")),
        ("D6", "ABAS",       str(len(resumo_categorias))),
    ]

    for addr, label, valor in metricas_cons:
        col_letter = addr[0]
        c = ws[addr]
        c.value     = valor
        c.font      = Font(name="Calibri", bold=True, size=20, color=t["metric_fg"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill      = _fill("FFFFFF")
        c.border    = Border(
            left  =Side(border_style="medium", color=t["card_border"]),
            right =Side(border_style="medium", color=t["card_border"]),
            top   =Side(border_style="medium", color=t["card_border"]),
            bottom=Side(border_style="medium", color=t["card_border"]),
        )
        label_c = ws.cell(5, ord(col_letter) - ord("A") + 1)
        label_c.value     = label
        label_c.font      = Font(name="Calibri", size=9, color=t["label_fg"], italic=True)
        label_c.alignment = Alignment(horizontal="center", vertical="bottom")
        label_c.fill      = fill_panel

    cabs = ["", "Nome da Aba", "Registros", "Arquivo de Origem"]
    for i, h in enumerate(cabs, start=1):
        cell = ws.cell(8, i)
        cell.value     = h
        cell.fill      = fill_header
        cell.font      = Font(name="Calibri", bold=True, size=10, color=t["header_fg"])
        cell.alignment = Alignment(
            horizontal="left" if i == 2 else "center",
            vertical="center",
            indent=1 if i == 2 else 0
        )
        cell.border = Border(bottom=Side(border_style="medium", color=t["header_bg"]))

    for i, (nome_aba, info) in enumerate(resumo_categorias.items(), start=9):
        is_par   = i % 2 == 0
        row_fill = fill_zebra if is_par else _fill("FFFFFF")
        vals = ["", nome_aba, f"{info['linhas']:,}".replace(",", "."),
                info["arquivos"][0] if info["arquivos"] else ""]
        aligns = ["center", "left", "center", "left"]
        indents = [0, 1, 0, 1]

        for j, (v, ha, ind) in enumerate(zip(vals, aligns, indents), start=1):
            cell = ws.cell(i, j)
            cell.value     = v
            cell.fill      = row_fill
            cell.font      = Font(name="Calibri", size=10, color="1C2B3A")
            cell.alignment = Alignment(horizontal=ha, vertical="center", indent=ind)
            cell.border    = Border(bottom=Side(border_style="thin", color=t["border"]))
        ws.row_dimensions[i].height = 26


def _noop(*args, **kwargs):
    pass


def processar_arquivo(caminho: str, tema: str = "🔵 Azul Executivo", ordenar: bool = True,
                      cb_prog=None, cb_log=None) -> str:
    cb_prog = cb_prog or _noop
    cb_log  = cb_log  or _noop

    caminho_saida = os.path.splitext(caminho)[0] + "_formatado.xlsx"
    cb_log("📂 Analisando arquivo principal...")
    cb_prog(5)

    wb = None
    try:
        try:
            wb = load_workbook(caminho, data_only=False)
        except InvalidFileException:
            raise Exception("Arquivo Excel inválido ou corrompido.")
        except FileNotFoundError:
            raise Exception("Arquivo não encontrado no sistema.")

        ws_inc = obter_ou_criar_aba_inconsistencias(wb)
        abas_processaveis = [n for n in wb.sheetnames if n != "⚠ Inconsistências"]
        total_invalidos, total_corrigidos = 0, 0

        for i, nome in enumerate(abas_processaveis, 1):
            ws_atual = wb[nome]
            cb_log(f"🧹 Removendo colunas técnicas de: {nome}")
            filtrar_colunas_util(ws_atual)

            cb_log(f"🧠 Validando aba: {nome} (Contexto: {detectar_modulo_por_aba(nome).title()})")
            resumo = validar_sheet(ws_atual, ws_inc)
            total_invalidos  += resumo["total_invalidos"]
            total_corrigidos += resumo["total_corrigidos"]
            cb_log(f"   ↳ Encontrados: {resumo['total_colunas_mapeadas']} colunas | "
                   f"{resumo['total_corrigidos']} corrigidos | {resumo['total_invalidos']} falhas")

            cb_log("📐 Ajustando largura das colunas...")
            auto_ajustar_largura(ws_atual)

            cb_log(f"🎨 Aplicando Design Premium em: {nome}")
            formatar_sheet(ws_atual, tema, ordenar)
            cb_prog(5 + int(75 * i / max(len(abas_processaveis), 1)))

        cb_log("🎨 Estruturando painel de inconsistências...")
        auto_ajustar_largura(ws_inc)
        formatar_sheet(ws_inc, tema, ordenar=False)
        cb_prog(85)

        cb_log("📋 Consolidando Dashboard...")
        criar_sumario(wb, tema)
        cb_prog(92)

        cb_log(f"📈 Resumo: {total_corrigidos} ajustes automáticos | {total_invalidos} falhas residuais")
        cb_log("💾 Fechando e gravando arquivo seguro...")

        try:
            wb.save(caminho_saida)
        except PermissionError:
            raise Exception(f"Feche a planilha no Excel antes de continuar:\n{caminho_saida}")

        cb_prog(100)
        cb_log("✅ OPERAÇÃO CONCLUÍDA.")
        return caminho_saida
    finally:
        if wb:
            try: wb.close()
            except Exception: pass


def processar_consolidacao(caminhos: list, tema: str = "🔵 Azul Executivo", ordenar: bool = True,
                           cb_prog=None, cb_log=None) -> str:
    cb_prog = cb_prog or _noop
    cb_log  = cb_log  or _noop

    wb_final    = Workbook()
    total_linhas = 0
    resumo       = {}

    try:
        for idx, caminho in enumerate(caminhos, 1):
            nome_arq = os.path.basename(caminho)
            cb_log(f"📂 Extraindo dados de: {nome_arq}")
            try:
                wb_origem = load_workbook(caminho, data_only=True, read_only=True)
            except Exception as e:
                cb_log(f"⚠️ Alerta em '{nome_arq}': {str(e)} - Arquivo ignorado.")
                continue
            try:
                for aba_origem in wb_origem.sheetnames:
                    nome_base = (f"{os.path.splitext(nome_arq)[0]}_{aba_origem}"
                                 if len(wb_origem.sheetnames) > 1
                                 else os.path.splitext(nome_arq)[0])
                    nome_aba  = limpar_nome_aba(nome_base, wb_final.sheetnames)
                    ws_nova   = wb_final.create_sheet(nome_aba)
                    for i, linha in enumerate(
                        (l for l in wb_origem[aba_origem].iter_rows(values_only=True)
                         if not linha_vazia(l)), 1
                    ):
                        copiar_linha(ws_nova, i, linha)
                        total_linhas += 1
                    filtrar_colunas_util(ws_nova)
                    resumo[nome_aba] = {"arquivos": [nome_arq], "linhas": ws_nova.max_row}
            finally:
                wb_origem.close()
            cb_prog(10 + int(45 * idx / len(caminhos)))

        if not resumo:
            raise Exception("Nenhum dado compatível localizado nos arquivos.")
        if "Sheet" in wb_final.sheetnames and len(wb_final.sheetnames) > 1:
            del wb_final["Sheet"]

        cb_log("🧾 Compilando Visão Geral...")
        cb_prog(65)
        criar_resumo_consolidacao(wb_final, caminhos, total_linhas, resumo, tema)
        ws_inc = obter_ou_criar_aba_inconsistencias(wb_final)

        cb_log("🧠 Iniciando validação especialista em lote...")
        abas = [a for a in wb_final.sheetnames
                if a not in ("📊 Dashboard", "📊 Visão Geral", "⚠ Inconsistências")]
        tot_inv, tot_corr = 0, 0

        for i, nome in enumerate(abas, 1):
            cb_log(f"🧠 Validando aba: {nome} (Contexto: {detectar_modulo_por_aba(nome).title()})")
            r_val = validar_sheet(wb_final[nome], ws_inc)
            tot_inv  += r_val["total_invalidos"]
            tot_corr += r_val["total_corrigidos"]
            cb_log(f"   ↳ {nome}: {r_val['total_corrigidos']} corretos | {r_val['total_invalidos']} falhas")

        cb_prog(78)
        cb_log("🎨 Aplicando Engenharia Visual e Ajuste de Dimensões...")
        for i, nome in enumerate(abas, 1):
            auto_ajustar_largura(wb_final[nome])
            formatar_sheet(wb_final[nome], tema, ordenar)
            cb_prog(78 + int(14 * i / max(len(abas), 1)))

        auto_ajustar_largura(ws_inc)
        formatar_sheet(ws_inc, tema, ordenar=False)

        cb_log("📋 Otimizando Interface do Arquivo...")
        cb_prog(93)
        cb_log(f"📈 Status Final: {tot_corr} ajustes | {tot_inv} bloqueados")

        saida = os.path.join(os.path.dirname(caminhos[0]), "consolidado_formatado.xlsx")
        cb_log("💾 Compilando arquivo seguro...")
        cb_prog(97)
        try:
            wb_final.save(saida)
        except PermissionError:
            raise Exception(f"Feche a planilha no Excel antes de continuar:\n{saida}")

        cb_prog(100)
        cb_log("✅ CONSOLIDAÇÃO CONCLUÍDA.")
        return saida
    finally:
        if wb_final:
            try: wb_final.close()
            except Exception: pass


def processar_arquivos(lista_arquivos, pasta_saida, tema="🔵 Azul Executivo", ordenar=True):
    if not lista_arquivos:
        raise Exception("Nenhum arquivo recebido para processamento.")

    pasta_saida = Path(pasta_saida)
    pasta_saida.mkdir(parents=True, exist_ok=True)

    if len(lista_arquivos) == 1:
        arquivo_gerado = processar_arquivo(lista_arquivos[0], tema=tema, ordenar=ordenar)
    else:
        arquivo_gerado = processar_consolidacao(lista_arquivos, tema=tema, ordenar=ordenar)

    origem  = Path(arquivo_gerado)
    destino = pasta_saida / origem.name

    if origem.resolve() != destino.resolve():
        shutil.copy2(origem, destino)

    return str(destino)